import torch
import torch.nn as nn
import torch.nn.functional as F
import torch.optim as optim
import pathlib
import openpyxl
import numpy as np
import matplotlib.pyplot as plt

# Import results
Read_file = pathlib.Path('D:\Vibration_control_Paper\Ansys_Code\Opt_Fitn.xlsx')
Output = openpyxl.load_workbook(Read_file)
Sheet = Output.active

Sheet_Alph = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'J']
Data_Fit = np.zeros((int(Sheet.max_row), int(Sheet.max_column-1)))
ctr = 0
for i in range(Sheet.max_row-4):
    ctr += 1
    for j in range(Sheet.max_column-1):
        Data_Fit[i, j] = Sheet[Sheet_Alph[j] + str(ctr)].value

TrainRatio = int(np.size(Data_Fit[1:2996, :], 0)*75/100)
#idx = np.sort(Data_Fit[0:TrainRatio, :], axis=1)
Train = Data_Fit[0:TrainRatio, :]
Test = Data_Fit[TrainRatio:2996, :]
Input1 = 8
Output1 = 8
Input2 = 8
Output2 = 8
Input3 = 8
Output3 = 8
Input4 = 8
Output4 = 1
Epoc_number = 10000
class N_Opt(nn.Module):

    def __init__(self):
        super().__init__()
        self.fc1 = nn.Linear(Input1, Output1)
        self.fc2 = nn.Linear(Input2, Output2)
        self.fc3 = nn.Linear(Input2, Output2)
        self.fc4 = nn.Linear(Input2, Output4)


    def forward(self, Out):
        Out = F.relu(self.fc1(Out))
        OUt = F.relu(self.fc2(Out))
        Out = F.relu(self.fc3(Out))

        Out = self.fc4(Out)
        return F.relu(Out)

net = N_Opt()

loss_function = nn.L1Loss()
Optimizer = optim.Adam(net.parameters(), lr=0.00092)
ctr = 0
erros = []
for epoch in range(Epoc_number):
    #for data in range(int(np.size(Train, 0))):
    y = Train[:, -1] / (np.max(Train[:, -1]))
    idx = np.argsort(y)
    #y = y[idx]
    X = Train[:, 0:8]
    #X = X[idx, :]

    Y = torch.from_numpy(y)#.type(torch.LongTensor)

    #print(X)
    #print(Y)
    net.zero_grad()
    output = net(torch.from_numpy(X.astype(np.float32)))
    loss = F.l1_loss(output, Y.unsqueeze(-1))
    loss.backward()
    Optimizer.step()
    #print(net.fc1.weight.grad)
    #print(net.fc2.weight.grad)
    ctr += 1
    print('epoch'+str(epoch), loss)
    v = loss.item()
    #erros = erros.append(float(v))


    plt.scatter(epoch, v)

correct = 0
total = 0

with torch.no_grad():
    #for data in range(int(np.size(Test, 0))):  # `data` is a batch of data
    X1 = Test[:, 0:8]
    y1 = Test[:, -1]/(np.max(Test[:, -1]))
    idx1 = np.argsort(y1)
    #y1 = y1[idx1]
    #X = X[idx1, :]

    Y1 = torch.from_numpy(y1)#.type(torch.LongTensor)
    output1 = net(torch.from_numpy(X1.astype(np.float32)))
    #nt(output1)
    for idx, i in enumerate(output1):
            #print(torch.argmax(i), y[idx])
        if torch.argmax(i) == Y1[idx]:
            correct += 1
        total += 1

    print("Accuracy: ", round(correct/total, 3))


plt.figure(2)
plt.plot(Y1.numpy() * np.max(Test[:, -1]), '.')
plt.plot((output1.numpy()) * np.max(Test[:, -1]), '.')

